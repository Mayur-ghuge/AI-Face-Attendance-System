from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

def mark_attendance(name):

    file = "attendance.xlsx"

    if not os.path.exists(file):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Date", "Time", "Status"])
        wb.save(file)

    wb = load_workbook(file)
    ws = wb.active

    today = datetime.now().strftime("%d/%m/%Y")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == name and row[1] == today:
            return

    now = datetime.now()

    ws.append([
        name,
        now.strftime("%d/%m/%Y"),
        now.strftime("%H:%M:%S"),
        "Present"
    ])

    wb.save(file)