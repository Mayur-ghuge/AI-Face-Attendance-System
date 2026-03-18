import customtkinter as ctk
import subprocess
import sys
import os
from datetime import datetime
from openpyxl import load_workbook
from PIL import Image
import threading

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

app = ctk.CTk()
app.title("FACE AI - Attendance System")
app.state("zoomed")

app.configure(fg_color="#0f172a")

# =========================
# 🔥 Preload DeepFace (better warmup)
# =========================

def preload_model():
    from deepface import DeepFace
    try:
        DeepFace.build_model("Facenet")
    except:
        pass

threading.Thread(target=preload_model, daemon=True).start()

# =========================
# 🔥 Loading Popup
# =========================

def show_loading(msg="Loading..."):
    loading = ctk.CTkToplevel(app)
    loading.geometry("300x120")
    loading.title("Please Wait")
    loading.grab_set()

    label = ctk.CTkLabel(loading, text=msg, font=("Arial",18))
    label.pack(expand=True)

    loading.update()
    return loading

# =========================
# Sidebar
# =========================

sidebar = ctk.CTkFrame(app, width=260, corner_radius=0, fg_color="#020617")
sidebar.pack(side="left", fill="y")

ctk.CTkLabel(sidebar, text="FACE AI", font=("Arial",30,"bold")).pack(pady=(30,10))

if os.path.exists("profile.png"):
    img = ctk.CTkImage(Image.open("profile.png"), size=(120,120))
    ctk.CTkLabel(sidebar, image=img, text="").pack(pady=10)

ctk.CTkLabel(sidebar, text="Mayur", font=("Arial",18,"bold")).pack(pady=(5,30))

# =========================
# 🚀 FIXED BUTTON FUNCTIONS
# =========================

def capture_dataset():
    loading = show_loading("Opening Camera...")
    app.after(100, lambda: subprocess.Popen([sys.executable,"dataset_capture.py"]))
    app.after(2000, loading.destroy)


# 🔥 FIX 1 (FAST RECOGNITION)
def start_recognition():
    loading = show_loading("Starting Face Recognition...")

    def run():
        os.system(f"{sys.executable} face_recognition.py")

    threading.Thread(target=run, daemon=True).start()

    app.after(2000, loading.destroy)


# 🔥 FIX 2 (FAST EXCEL OPEN)
def view_attendance():
    if os.path.exists("attendance.xlsx"):
        threading.Thread(
            target=lambda: os.system("open attendance.xlsx"),
            daemon=True
        ).start()


def exit_app():
    app.destroy()

# Buttons
ctk.CTkButton(sidebar,text="📸 Capture Dataset",command=capture_dataset,width=200,height=45).pack(pady=10)
ctk.CTkButton(sidebar,text="🧠 Start Recognition",command=start_recognition,width=200,height=45).pack(pady=10)
ctk.CTkButton(sidebar,text="📊 View Attendance",command=view_attendance,width=200,height=45).pack(pady=10)
ctk.CTkButton(sidebar,text="🚪 Exit",command=exit_app,fg_color="red",hover_color="#8b0000",width=200,height=45).pack(pady=30)

# =========================
# Main Dashboard
# =========================

main = ctk.CTkFrame(app, fg_color="#0f172a")
main.pack(side="right", expand=True, fill="both", padx=40, pady=40)

ctk.CTkLabel(main,text="Dashboard",font=("Arial",40,"bold")).pack(pady=10)

# Time
time_label = ctk.CTkLabel(main,font=("Arial",18))
time_label.pack()

def update_time():
    now = datetime.now().strftime("%A, %d %B %Y | %H:%M:%S")
    time_label.configure(text=now)
    app.after(1000, update_time)

update_time()

# =========================
# Stats
# =========================

stats_frame = ctk.CTkFrame(main, fg_color="#0f172a")
stats_frame.pack(pady=50)

cards = []
labels = ["Total Students","Present Today","Absent Today"]

for i in range(3):
    card = ctk.CTkFrame(stats_frame,width=260,height=150,fg_color="#1e293b")
    card.grid(row=0,column=i,padx=25)
    
    ctk.CTkLabel(card,text=labels[i],font=("Arial",18)).pack(pady=10)
    val = ctk.CTkLabel(card,font=("Arial",36,"bold"))
    val.pack()
    cards.append(val)

total_value, present_value, absent_value = cards

# =========================
# Attendance Table
# =========================

table_frame = ctk.CTkFrame(main, fg_color="#1e293b")
table_frame.pack(pady=20)

ctk.CTkLabel(table_frame,text="Today's Attendance",font=("Arial",22,"bold")).pack(pady=10)

table_box = ctk.CTkTextbox(table_frame,width=750,height=230)
table_box.pack()

DATASET = "images/student_photos"
ATTENDANCE = "attendance.xlsx"

# =========================
# Update Stats
# =========================

def update_stats():

    total_students = len(os.listdir(DATASET)) if os.path.exists(DATASET) else 0
    present_today = 0
    today = datetime.now().strftime("%d/%m/%Y")

    table_box.delete("1.0","end")

    if os.path.exists(ATTENDANCE):
        wb = load_workbook(ATTENDANCE)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            name, date, time = row[:3]

            if date == today:
                present_today += 1
                table_box.insert("end", f"{name}   |   {time}\n")

    absent_today = max(total_students - present_today, 0)

    total_value.configure(text=total_students)
    present_value.configure(text=present_today)
    absent_value.configure(text=absent_today)

update_stats()

app.mainloop()